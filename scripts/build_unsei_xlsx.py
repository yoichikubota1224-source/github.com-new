#!/usr/bin/env python3
# 騎手運勢xlsxに 2026.08.29 シート(8/29土・8/30日)を先頭追加する。
# r3の教訓に従い、セルの塗りは「値から再構築」せず**原本セルの書式オブジェクトを複製**して適用する。
import copy, datetime, os, sys
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from gen_unsei import unsei, GROUPS, serial

SRC = '/home/user/github.com-new/predictions/20260823/騎手運勢2026.08.22.xlsx'
DST = os.path.join(os.path.dirname(os.path.abspath(__file__)), '騎手運勢2026.08.29.xlsx')
BASE_SHEET = '2026.08.22'
NEW_SHEET = '2026.08.29'
D1, D2 = datetime.date(2026, 8, 29), datetime.date(2026, 8, 30)

wb = openpyxl.load_workbook(SRC)
base = wb[BASE_SHEET]

# 記号 -> 原本セル(書式の複製元)。48セルすべてで書式が一意であることは検証済み。
proto = {}
for r in range(3, 27):
    for c in (2, 3):
        proto.setdefault(base.cell(r, c).value, base.cell(r, c))
assert set(proto) == {'◎◎', '◎', '○', '△', '×'}, proto.keys()

new = wb.copy_worksheet(base)
new.title = NEW_SHEET
wb.move_sheet(NEW_SHEET, offset=-(wb.sheetnames.index(NEW_SHEET)))   # 先頭へ

new['B2'] = serial(D1)
new['C2'] = serial(D2)

for i, g in enumerate(GROUPS):
    row = 3 + i
    assert new.cell(row, 1).value == g, (row, new.cell(row, 1).value, g)
    for col, d in ((2, D1), (3, D2)):
        cyc, mark = unsei(d, g)
        cell = new.cell(row, col)
        cell.value = mark
        p = proto[mark]
        cell.fill = copy.copy(p.fill)       # 原本オブジェクトの複製(RGB/theme/tint/end_colorごと)
        cell.font = copy.copy(p.font)
        cell.alignment = copy.copy(p.alignment)
        cell.border = copy.copy(p.border)

wb.save(DST)
print('saved', DST, os.path.getsize(DST), 'bytes /', len(wb.sheetnames), 'sheets')
